import os
import sys
import logging
from datetime import datetime, timedelta
import pandas as pd
import xlwings as xw
from xlwings.constants import AutoFillType
import calendar
 
import Infinox_Sales_SQL
 
sys.path.append(r"\\192.168.1.20\Rmc\Data Analysis\Dropbox\Model_AUTO")
from package.date import dateT, datecollall
#from package.skype import skp_send
 
 
# Skype 連線資訊
#pass_dir = r"\\192.168.1.20\Rmc\Data Analysis\Dropbox\Model_AUTO\log\pass\pass.xlsx"
#login = str(pd.read_excel(pass_dir).query("item=='skype'")["userid"][0])
#pwd = pd.read_excel(pass_dir).query("item=='skype'")["pwd"][0]
 
# log 生成位置及檔案名稱
log_dir = r"\\Users\beatrice.law\Desktop\Daily Report - B\Infinox Weekly Office Gross Deposit"
log_file = "Infinox_weekly_office.log"
 
 
Rundate = datecollall(dateT()[0],dateT()[0])
# Rundate = datecollall("2023-07-19","2023-07-19")
# num = Rundate[0]
 
def get_last_month_mondays_and_sundays():
    # Get the current date
    today = datetime.today()
 
    # Get the first day of the current month
    first_day_of_current_month = today.replace(day=1)
 
    # Calculate the first day of the last month
    first_day_of_last_month = first_day_of_current_month - timedelta(days=first_day_of_current_month.day)
 
    # Extract the year and month of the last month
    year = first_day_of_last_month.year
    month = first_day_of_last_month.month
   
    # Get the calendar of the last month
    cal = calendar.monthcalendar(year, month)
 
    # Check if the first day of the last month is a weekend (Saturday/Sunday)
    if cal[0][calendar.SATURDAY] != 0 or cal[0][calendar.SUNDAY] != 0:
        # If the first day is a weekend, skip the first Monday
        mondays = [week[calendar.MONDAY] for week in cal[1:] if week[calendar.MONDAY] != 0]
    else:
        # Otherwise, include all Mondays
        mondays = [week[calendar.MONDAY] for week in cal if week[calendar.MONDAY] != 0]
 
    # Check if the first day of the last month is a Sunday
    if cal[0][calendar.SUNDAY] != 0:
        # If the first day is a Sunday, skip the first Sunday
        sundays = [week[calendar.SUNDAY] for week in cal[1:] if week[calendar.SUNDAY] != 0]
    else:
        # Otherwise, include all Sundays
        sundays = [week[calendar.SUNDAY] for week in cal if week[calendar.SUNDAY] != 0]
 
    # Check if the last day of the last month is a Sunday (index -1)
    last_day = cal[-1][calendar.SUNDAY]
    days_in_month = calendar.monthrange(year, month)[1]
    if last_day != days_in_month:
        # If the last day isn't a Sunday, append it as the last entry in the Sundays list
        sundays.append(days_in_month)
 
    # Construct the dates of Mondays and Sundays in the last month
    mondays_dates = [f"{year}-{month:02d}-{day:02d}" for day in mondays if day != 0]
    sundays_dates = [f"{year}-{month:02d}-{day:02d}" for day in sundays if day != 0]
 
    return mondays_dates, sundays_dates, month
 
print(month)
 
# Get the list of Mondays' and Sundays' dates in the last month
mondays, sundays = get_last_month_mondays_and_sundays()
 
try:
    for num in Rundate:
        print(f'start doing {num[0]} Infinox Weeekly Office Gross Deposit!')
 
        # 計算運行時間
        begin = datetime.now()
       
        # catch SQL data
        detail_raw = Infinox_Sales_SQL.detail_raw(num)
        regulation = Infinox_Sales_SQL.regulation(num)
        anguilla = Infinox_Sales_SQL.anguilla(num)
 
 
        infinox_weekly_office_model_path = fr"\Users\beatrice.law\Desktop\Daily Report - B\Infinox Weekly Office Gross Deposit"
        infinox_weekly_office_path = fr"\Users\beatrice.law\Desktop\Daily Report - B\Infinox Weekly Office Gross Deposit"
 
 
        # open Infinox Weekly Office Gross Profit
        infinox_weekly_office_model = f"{infinox_weekly_office_model_path}\Infinox Weekly Gross Deposit (Nov).xlsx"
        infinox_weekly_office = f"{infinox_weekly_office_path}\Infinox Weekly Gross Deposit {month}.xlsx"
        app = xw.App(visible=False, add_book=False)
        infinox = app.books.open(infinox_weekly_office_model)
 
        # 找出最後一列的index
        def last_row(sheetName, cells):
            return sheetName.range(cells).expand("down").last_cell.row
       
        # 解除篩選器
        for sheet in infinox.sheets:
            if sheet.api.AutoFilterMode == True:
                sheet.api.AutoFilterMode = False
 
        # setup all sheets
        Weekly_Summary_sht = infinox.sheets["Weekly Gross Deposit Summary"]
        Weekly_Result_sht = infinox.sheets["Weekly Gross Deposit Result"]
        Infinox_detail_sht = infinox.sheets["Infinox_detail"]
        Sales_Office_Data_sht = infinox.sheets["Sales&Office_Data"]
        Anguilla_sht = infinox.sheets["Anguilla"]
       
        ## Weekly Gross Deposit Summary
        Weekly_Summary_sht.range("Q6").options(header=False, index=False).value = mondays[0]
        Weekly_Summary_sht.range("R6").options(header=False, index=False).value = sundays[0]
        Weekly_Summary_sht.range("Q7").options(header=False, index=False).value = mondays[1]
        Weekly_Summary_sht.range("R7").options(header=False, index=False).value = sundays[1]
        Weekly_Summary_sht.range("Q8").options(header=False, index=False).value = mondays[2]
        Weekly_Summary_sht.range("R8").options(header=False, index=False).value = sundays[2]
        Weekly_Summary_sht.range("Q9").options(header=False, index=False).value = mondays[3]
        Weekly_Summary_sht.range("R9").options(header=False, index=False).value = sundays[3]
        #Weekly_Summary_sht.range("Q10").options(header=False, index=False).value = mondays[4]
        #Weekly_Summary_sht.range("R10").options(header=False, index=False).value = sundays[4]
 
 
        ## Weekly Gross Deposit Result
        Weekly_Result_sht.range("P6").options(header=False, index=False).value = mondays[0]
        Weekly_Result_sht.range("Q6").options(header=False, index=False).value = sundays[0]
        Weekly_Result_sht.range("P7").options(header=False, index=False).value = mondays[1]
        Weekly_Result_sht.range("Q7").options(header=False, index=False).value = sundays[1]
        Weekly_Result_sht.range("P8").options(header=False, index=False).value = mondays[2]
        Weekly_Result_sht.range("Q8").options(header=False, index=False).value = sundays[2]
        Weekly_Result_sht.range("P9").options(header=False, index=False).value = mondays[3]
        Weekly_Result_sht.range("Q9").options(header=False, index=False).value = sundays[3]
        #Weekly_Result_sht.range("P10").options(header=False, index=False).value = mondays[4]
        #Weekly_Result_sht.range("Q10").options(header=False, index=False).value = sundays[4]
       
       
        ## Infinox_detail
        # clear & paste
        Infinox_detail_sht.range("A3:AE{}".format(last_row(Infinox_detail_sht, "A3"))).clear_contents()
        Infinox_detail_sht.range("A2").options(header=False, index=False).value = detail_raw
        # adjust formula
        if len(detail_raw) > 1:
            Infinox_detail_sht.range("V2:AE2").api.AutoFill(
                Infinox_detail_sht.range("V2:AE{}".format(last_row(Infinox_detail_sht, "A2"))).api, AutoFillType.xlFillDefault)
        print("Infinox_detail Done")
 
 
        ## Sales&Office_Data
        # clear & paste
        Sales_Office_Data_sht.range("A3:I{}".format(last_row(Sales_Office_Data_sht, "A3"))).clear_contents()
        Sales_Office_Data_sht.range("A2").options(header=False, index=False).value = regulation
        # adjust formula
        if len(regulation) > 1:
            Sales_Office_Data_sht.range("H2:I2").api.AutoFill(
                Sales_Office_Data_sht.range("H2:I{}".format(last_row(Sales_Office_Data_sht, "A2"))).api, AutoFillType.xlFillDefault)
        print("Sales&Office_Data Done")
 
 
        ## Anguilla
        # clear & paste
        Anguilla_sht.range("A3:E{}".format(last_row(Anguilla_sht, "B3"))).clear_contents()
        Anguilla_sht.range("B2").options(header=False, index=False).value = anguilla
        # adjust formula
        if len(anguilla) > 1:
            Anguilla_sht.range("A2").api.AutoFill(
                Anguilla_sht.range("A2:A{}".format(last_row(Anguilla_sht, "B2"))).api, AutoFillType.xlFillDefault)
            Anguilla_sht.range("E2").api.AutoFill(
                Anguilla_sht.range("E2:E{}".format(last_row(Anguilla_sht, "B2"))).api, AutoFillType.xlFillDefault)
        print("Anguilla Done")
       
 
        # 關閉 Infinox Sales, Excel app
        infinox.save(infinox_weekly_office)
        infinox.close()
        app.quit()
       
        finish = datetime.now()
        print(f"{num[0]} Infinox Weekly Office Gross Deposit Done")
 
except:
    app.quit()
    logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s", filename=os.path.join(log_dir, log_file))
    logging.critical("Catch an exception.", exc_info=True)
    #skp_send(login, pwd, who="auto_fail", content=f"{num[0]} {log_file} fail!",
             #file_dir=log_dir, file=log_file, tag='Erica')
    logging.shutdown()
    os.remove(os.path.join(log_dir,log_file))
    raise SystemExit("try again!")
 
#else:
    #skp_send(login, pwd, who="auto_fail", content=f"{num[0]} Infinox Sales Done (spent : {finish-begin})")
 
from datetime import datetime, timedelta
from openpyxl import Workbook
 
def get_previous_month():
    today = datetime.today()
    first_day_previous_month = datetime(today.year, today.month, 1) - timedelta(days=1)
    year = first_day_previous_month.year
    month = first_day_previous_month.month
   
    # Set date
    # year = 2024
    # month = 12
    return year, month
 
def get_sunday_of_week(date):
    return date - timedelta(days=date.weekday())
 
def get_weekly_intervals(year, month):
    # Find the first day of the month
    first_day_of_month = datetime(year, month, 1)
 
    # Find the first Sunday of the month
    first_sunday = get_sunday_of_week(first_day_of_month)
 
    # Check if the next month exceeds 12, adjust year and set month to 1
    next_month = month + 1
    if next_month > 12:
        next_month = 1
        year += 1
 
    last_day_of_month = datetime(year, next_month, 1) - timedelta(days=1)
 
    weekly_intervals = []
    current_day = first_sunday
 
    while current_day <= last_day_of_month:
        current_saturday = min(current_day + timedelta(days=6), last_day_of_month)
       
        # Adjust start day to the first day of the month for Week 1
        start_day = first_day_of_month if current_day == first_sunday else current_day
 
        weekly_intervals.append({
            "Week": len(weekly_intervals) + 1,
            "From": start_day,
            "To": current_saturday
        })
        current_day = current_saturday + timedelta(days=1)
 
    return weekly_intervals
 
from openpyxl import load_workbook
 
existing_filename = infinox_weekly_office
def write_to_existing_excel(weekly_intervals, existing_filename):
    try:
        # Load existing workbook
        wb = load_workbook(existing_filename)
 
        # Get existing worksheets or create new ones if they don't exist
        ws_summary = wb.get_sheet_by_name("Weekly Gross Deposit Summary") if "Weekly Gross Deposit Summary" in wb.sheetnames else wb.create_sheet("Weekly Gross Deposit Summary")
        ws_result = wb.get_sheet_by_name("Weekly Gross Deposit Result") if "Weekly Gross Deposit Result" in wb.sheetnames else wb.create_sheet("Weekly Gross Deposit Result")
 
        # Specify the starting row and column for the new data
        start_row = 5
        start_col = 15  # Column O
 
        # Write data for both worksheets
        for i, interval in enumerate(weekly_intervals, start=1):
            # Exclude weeks where From and To are empty
            if interval["From"] and interval["To"]:
                ws_summary.cell(row=i + start_row, column=start_col, value=f"Week {interval['Week']}")
                ws_summary.cell(row=i + start_row, column=start_col + 1, value=interval['From'].strftime("%m/%d/%Y"))
                ws_summary.cell(row=i + start_row, column=start_col + 2, value=interval['To'].strftime("%m/%d/%Y"))
 
                ws_result.cell(row=i + start_row, column=start_col, value=f"Week {interval['Week']}")
                ws_result.cell(row=i + start_row, column=start_col + 1, value=interval['From'].strftime("%m/%d/%Y"))
                ws_result.cell(row=i + start_row, column=start_col + 2, value=interval['To'].strftime("%m/%d/%Y"))
 
        # Save the workbook
        wb.save(existing_filename)
        print(f"Data appended to {existing_filename}")
 
    except Exception as e:
        print(f"Error: {e}")
 
 
# Get weekly intervals
previous_year, previous_month = get_previous_month()
weekly_intervals = get_weekly_intervals(previous_year, previous_month)
 
# Append data to the specified range in the existing workbook
write_to_existing_excel(weekly_intervals, existing_filename)