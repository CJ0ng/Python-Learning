import openpyxl
from openpyxl.chart import BarChart, PieChart3D, Reference
from datetime import datetime

def modify_data(sheet, min_col, max_col, cap_value=None, exclude_last_row=False):
    end_row = sheet.max_row - 1 if exclude_last_row else sheet.max_row
    for row in sheet.iter_rows(min_row=2, max_row=end_row, min_col=min_col, max_col=max_col):
        for cell in row:
            if cell.value is not None and cap_value is not None:
                cell.value = min(cell.value, cap_value)

def create_and_save_chart(file_path, sheet_name, create_chart_func, title, min_col, max_col, chart_location, cap_value=None, exclude_last_row=False):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(title=sheet_name)
    modify_data(sheet, min_col, max_col, cap_value, exclude_last_row)
    chart = create_chart_func()
    chart.title = title
    end_row = sheet.max_row - 1 if exclude_last_row else sheet.max_row
    chart.add_data(Reference(sheet, min_col=min_col, min_row=1, max_col=max_col, max_row=end_row), titles_from_data=True)
    chart.set_categories(Reference(sheet, min_col=1, min_row=2, max_row=sheet.max_row))
    chart.width, chart.height = 15, 10
    sheet.add_chart(chart, chart_location)
    wb.save(file_path)

# Monthly Slippage for FCA
create_and_save_chart(
    r"C:\Users\chungjing.ong\Documents\Python WorkSpace\Python Learning\Training\TestChart2.xlsx",
    "Monthly_Slippage_FCA",
    BarChart,
    "Slippage - Monthly",
    2, 4,
    "F2"
)

# Slippage Report by LP for FCA
create_and_save_chart(
    r"C:\Users\chungjing.ong\Documents\Python WorkSpace\Python Learning\Training\TestChart2.xlsx",
    "Slippage_LP_FCA",
    BarChart,
    f"Slippage Report by LP in {datetime.now().strftime('%B')}",
    6, 8,
    "J3",
    cap_value=120,
    exclude_last_row=True
)

# Slippage Percentage for FCA
create_and_save_chart(
    r"C:\Users\chungjing.ong\Documents\Python WorkSpace\Python Learning\Training\TestChart2.xlsx",
    "Slippage_Percentage_FCA",
    PieChart3D,
    f"Slippage in {datetime.now().strftime('%B')}",
    4, 4,
    "F3",
    cap_value=100
)

# Monthly Slippage for FSC
create_and_save_chart(
    r"C:\Users\chungjing.ong\Documents\Python WorkSpace\Python Learning\Training\TestChart.xlsx",
    "Monthly_Slippage_FSC",
    BarChart,
    "Slippage - Monthly",
    2, 4,
    "F2"
)

# Slippage Report by LP for FSC
create_and_save_chart(
    r"C:\Users\chungjing.ong\Documents\Python WorkSpace\Python Learning\Training\TestChart.xlsx",
    "Slippage_LP_FSC",
    BarChart,
    f"Slippage Report by LP in {datetime.now().strftime('%B')}",
    6, 8,
    "J3",
    cap_value=120,
    exclude_last_row=True
)

# Slippage Percentage for FSC
create_and_save_chart(
    r"C:\Users\chungjing.ong\Documents\Python WorkSpace\Python Learning\Training\TestChart.xlsx",
    "Slippage_Percentage_FSC",
    PieChart3D,
    f"Slippage in {datetime.now().strftime('%B')}",
    4, 4,
    "F3",
    cap_value=100
)
